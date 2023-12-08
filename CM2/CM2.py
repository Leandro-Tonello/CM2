import re
import openpyxl
import pandas as pd
import datetime
import random
import string
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import ttk
from Clases.Tarjeta import Tarjeta
from Clases.Persona import Persona

def generar_archivo():
     try:
        # Obtener la ubicacion del archivo de entrada
        archivo_excel = entrada_archivo.get()
        if not archivo_excel:
            raise ValueError("No se selecciono un archivo de entrada.")

        # Lee el archivo Excel en un DataFrame de pandas
        df = pd.read_excel(archivo_excel, header=None)  # No asume que la primera fila es el encabezado

        # Creo el filtro que separa una persona de otra
        filas_apellido_nombre = df[df.eq('Apellido y Nombre').any(axis=1)].index

        #Inicializa una lista para almacenar todas las instancias de Persona
        todas_las_personas = []

        # Itera sobre las filas con "Apellido y Nombre" y crea instancias de Persona
        for fila in filas_apellido_nombre:
            nombre = df.loc[fila + 1, 1]  # Se asume que los nombres están en la segunda columna (columna 1)
            dni_raw = df.loc[fila + 1, 2]  # Se asume que los DNI están en la tercera columna (columna 2)
  
            # Busquedas desde la celda Celular
            fila_celular = df[df.eq('Celular').any(axis=1)].index[df[df.eq('Celular').any(axis=1)].index > fila].min()
            celular = df.loc[fila_celular, 1]  # Se asume que los números de celular están en la segunda columna (columna 1)
            telefono = df.loc[fila_celular - 1, 1]  
            email = df.loc[fila_celular + 1, 1]  

            # Busquedas desde la celda Nro. Tarjeta CBU
            fila_tarjeta = df[df.eq('Nro. Tarjeta / CBU').any(axis=1)].index[df[df.eq('Nro. Tarjeta / CBU').any(axis=1)].index > fila].min()
            tarjeta1 = df.loc[fila_tarjeta + 1, 1]
            tarjeta2 = df.loc[fila_tarjeta + 2, 1]
            try:
                tarjeta3 = df.loc[fila_tarjeta + 3, 1]  # Accede a la celda que está tres filas debajo de "Tarjeta"
            except KeyError:
                tarjeta3 = None 
            try:
                tarjeta4 = df.loc[fila_tarjeta + 4, 1]  # Accede a la celda que está tres filas debajo de "Tarjeta"
            except KeyError:
                tarjeta4 = None

            if pd.notna(nombre) and pd.notna(dni_raw) and pd.notna(celular) and pd.notna(telefono) and pd.notna(email):
                dni = dni_raw.replace('DNI.', '').strip()
                persona = Persona(nombre, dni, celular, telefono, email, tarjeta1, tarjeta2, tarjeta3, tarjeta4)
                todas_las_personas.append(persona)



        # Crear un libro y una hoja de calculo con openpyxl
        workbook = Workbook()
        worksheet = workbook.active

        # Crear el DataFrame
        data = []
        for persona in todas_las_personas:
            data.append([
                persona.numero_persona,
                persona.nombre,
                persona.dni,
                persona.celular,
                persona.telefono,
                persona.email,
                "'"+persona.t1,
                persona.v1,
                "'"+persona.t2,
                persona.v2,
                "'"+persona.t3,
                persona.v3,
                "'"+persona.t4,
                persona.v4,
            ])

        df = pd.DataFrame(data, columns=['Nro', 'Nombre','Documento', 'Telefono-1', 'Telefono-2', 'Email','Tarjeta-1', 'Validado?', 'Tarjeta-2', 'Validado?', 'Tarjeta-3', 'Validado?', 'Tarjeta-4', 'Validado?',])

        # Escribir los datos en la hoja de calculo
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)

        # Configurar el formato y estilo
        header_fill = PatternFill(fill_type='solid', fgColor='000000')
        header_font = Font(color='FFFFFF', bold=True)
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        gray_fill = PatternFill(fill_type='solid', fgColor='E9E9E9')

        # Aplicar formato y estilo a los encabezados
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Aplicar formato y estilo a las celdas de datos
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=df.shape[1]):
            for cell in row:
                cell.border = data_border

        # Aplicar formato y estilo a las filas alternas
        for row_num in range(2, worksheet.max_row + 1, 2):
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = gray_fill

        # Ajustar el ancho de las columnas
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Generar un nombre de archivo aleatorio basado en la fecha y hora
        timestamp = datetime.datetime.now().strftime("%Y%m%d")
        random_word = ''.join(random.choices(string.ascii_lowercase, k=5))
        nombre_archivo = f"{timestamp}_{random_word}.xlsx"

        # Ruta de guardado con el nombre de archivo generado
        ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx")

        if not ruta_guardado:
            raise ValueError("No se selecciono una ubicacion de guardado.")

        # Guardar el archivo Excel
        workbook.save(ruta_guardado)
        workbook.close()

        # Mostrar la ubicacion del archivo generado
        resultado_archivo.config(text="Archivo generado:\n" + ruta_guardado)
        
     except ValueError as e:
      resultado_archivo.config(text=str(e))

# Funcion para seleccionar un archivo
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    entrada_archivo.delete(0, tk.END)
    entrada_archivo.insert(0, archivo)


# Crear la interfaz de usuario
window = tk.Tk()
window.title("Generador de Archivo Excel")

style = ttk.Style()
style.configure("TButton", font=("Arial", 12), padding=6)

# Etiqueta para mostrar el texto "Seleccione el Excel"
lbl_archivo = ttk.Label(window, text="Seleccione el Excel:", font=("Arial", 12, "bold"))
lbl_archivo.pack(pady=10)

# Marco que contiene el campo de entrada y el boton "Seleccionar"
frame_seleccionar = ttk.Frame(window)
frame_seleccionar.pack(pady=5)

# Campo de entrada deshabilitado para mostrar la ubicacion del archivo seleccionado
entrada_archivo = ttk.Entry(frame_seleccionar, width=50)
entrada_archivo.pack(side=tk.LEFT, padx=5)

# Boton "Seleccionar" para elegir un archivo
btn_seleccionar = ttk.Button(frame_seleccionar, text="Seleccionar", command=seleccionar_archivo)
btn_seleccionar.pack(side=tk.LEFT, padx=5)

# Boton "Generar" para generar el archivo Excel
btn_generar = ttk.Button(window, text="Generar", style="TButton", command=generar_archivo)
btn_generar.pack(pady=10)

# Etiqueta para mostrar la ubicacion del archivo generado
resultado_archivo = ttk.Label(window, text="", font=("Arial", 12))
resultado_archivo.pack()


window.mainloop()